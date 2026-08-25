import csv
import io
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import ImportedStatement, Movement, RecordState, ReconciliationState, StatementLine, StatementLineState

DEFAULT_TOLERANCE_DIAS = 1

_FECHA_ALIASES = {"fecha", "date", "fecha operacion", "fecha de la operacion", "fecha de origen"}
_MONTO_ALIASES = {"monto", "importe", "amount", "credito", "haber", "importe credito", "valor de la compra"}
_DESCRIPCION_ALIASES = {"descripcion", "concepto", "detalle", "description", "pagador"}
_REFERENCIA_ALIASES = {
    "referencia",
    "numero_operacion",
    "nro_operacion",
    "nro operacion",
    "comprobante",
    "numero de operacion",
    "id de operacion en mercado pago",
}

_DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y")


@dataclass
class StatementRow:
    fecha: datetime
    monto: Decimal
    descripcion: str | None
    referencia: str | None


class StatementParseError(ValueError):
    pass


def _normalize_header(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode().strip().lower()
    return text


def _match_column(headers: list[str], aliases: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if header in aliases:
            return index
    return None


def _parse_fecha(raw: str) -> datetime:
    raw = raw.strip()
    try:
        # Cubre timestamps ISO 8601 completos, con hora y zona horaria (ej. Mercado
        # Pago: "2026-08-21T18:16:29.000-04:00"); la tolerancia de matching es por
        # dia, asi que solo interesa la fecha, no la hora ni la zona horaria.
        parsed = datetime.fromisoformat(raw)
        return datetime(parsed.year, parsed.month, parsed.day)
    except ValueError:
        pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    raise StatementParseError(f"No se pudo interpretar la fecha '{raw}'.")


def _parse_monto(raw: str) -> Decimal:
    text = str(raw).strip().replace("$", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rindex(",") > text.rindex("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise StatementParseError(f"No se pudo interpretar el monto '{raw}'.") from exc


def _cell_or_none(row: list, index: int | None) -> str | None:
    if index is None or row[index] in (None, ""):
        return None
    return str(row[index]).strip()


def _rows_from_table(header_row: list, data_rows: list[list]) -> list[StatementRow]:
    headers = [_normalize_header(h) for h in header_row]
    idx_fecha = _match_column(headers, _FECHA_ALIASES)
    idx_monto = _match_column(headers, _MONTO_ALIASES)
    idx_descripcion = _match_column(headers, _DESCRIPCION_ALIASES)
    idx_referencia = _match_column(headers, _REFERENCIA_ALIASES)

    if idx_fecha is None or idx_monto is None:
        raise StatementParseError(
            "El archivo debe tener columnas de fecha y monto (por ejemplo 'Fecha' e 'Importe')."
        )

    rows: list[StatementRow] = []
    for raw_row in data_rows:
        if all(cell in (None, "") for cell in raw_row):
            continue
        fecha_cell = raw_row[idx_fecha]
        monto_cell = raw_row[idx_monto]
        if fecha_cell in (None, "") or monto_cell in (None, ""):
            continue
        fecha = fecha_cell if isinstance(fecha_cell, datetime) else _parse_fecha(str(fecha_cell))
        monto = _parse_monto(monto_cell)
        descripcion = _cell_or_none(raw_row, idx_descripcion)
        referencia = _cell_or_none(raw_row, idx_referencia)
        rows.append(StatementRow(fecha=fecha, monto=monto, descripcion=descripcion, referencia=referencia))
    return rows


def parse_csv(contenido: bytes) -> list[StatementRow]:
    text = contenido.decode("utf-8-sig")
    sample = text[:2048]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = list(reader)
    if not all_rows:
        raise StatementParseError("El archivo CSV esta vacio.")
    return _rows_from_table(all_rows[0], all_rows[1:])


def parse_xlsx(contenido: bytes) -> list[StatementRow]:
    workbook = load_workbook(io.BytesIO(contenido), data_only=True)
    sheet = workbook.active
    all_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    if not all_rows:
        raise StatementParseError("El archivo Excel esta vacio.")
    return _rows_from_table(all_rows[0], all_rows[1:])


def parse_statement_file(nombre_archivo: str, contenido: bytes) -> tuple[list[StatementRow], str]:
    lower = nombre_archivo.lower()
    if lower.endswith(".csv"):
        return parse_csv(contenido), "csv"
    if lower.endswith(".xlsx"):
        return parse_xlsx(contenido), "xlsx"
    raise StatementParseError("Formato no soportado todavia: usa CSV o XLSX.")


def _candidatos_iniciales(session: Session, cuenta_bancaria_id: int) -> list[Movement]:
    """Movimientos que podrian matchear contra algun resumen de esta cuenta, consultados
    una unica vez: por linea solo se filtra en memoria (ver match_statement)."""
    movimientos = session.scalars(
        select(Movement).where(
            Movement.estado_registro == RecordState.CONFIRMADO,
            Movement.estado_conciliacion.in_([ReconciliationState.PENDIENTE, ReconciliationState.CON_DIFERENCIA]),
        )
    ).all()
    return [
        m for m in movimientos if m.cuenta_bancaria_id is None or m.cuenta_bancaria_id == cuenta_bancaria_id
    ]


def _dentro_de_fecha(movimiento: Movement, linea: StatementLine, tolerancia_dias: int) -> bool:
    if movimiento.fecha_transaccion is None:
        return False
    return abs((movimiento.fecha_transaccion - linea.fecha).days) <= tolerancia_dias


def match_statement(
    session: Session,
    resumen: ImportedStatement,
    lineas: list[StatementLine],
    tolerancia_dias: int = DEFAULT_TOLERANCE_DIAS,
) -> None:
    usados: set[int] = set()
    pool = _candidatos_iniciales(session, resumen.cuenta_bancaria_id)

    for linea in lineas:
        candidatos = [m for m in pool if m.id not in usados]
        por_fecha = [m for m in candidatos if _dentro_de_fecha(m, linea, tolerancia_dias)]
        exactos = [m for m in por_fecha if m.monto == linea.monto]

        if linea.referencia:
            por_referencia = [m for m in exactos if m.numero_operacion == linea.referencia]
            if len(por_referencia) == 1:
                _asignar(linea, por_referencia[0], ReconciliationState.CONCILIADO, usados)
                continue

        if len(exactos) == 1:
            _asignar(linea, exactos[0], ReconciliationState.CONCILIADO, usados)
            continue
        if len(exactos) > 1:
            continue

        if linea.referencia:
            con_diferencia = [m for m in por_fecha if m.numero_operacion == linea.referencia]
            if len(con_diferencia) == 1:
                _asignar(linea, con_diferencia[0], ReconciliationState.CON_DIFERENCIA, usados)
                continue


def _asignar(
    linea: StatementLine, movimiento: Movement, estado: ReconciliationState, usados: set[int]
) -> None:
    linea.movimiento_id = movimiento.id
    linea.estado = StatementLineState.CONCILIADA
    movimiento.estado_conciliacion = estado
    if movimiento.cuenta_bancaria_id is None:
        movimiento.cuenta_bancaria_id = linea.resumen.cuenta_bancaria_id
    usados.add(movimiento.id)
